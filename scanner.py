import fitz  # PyMuPDF
import os, sys
import time
import random
import glob
import hashlib
import requests
import openpyxl
import pandas as pd
import threading
import tkinter as tk
from concurrent.futures import ThreadPoolExecutor, wait, ALL_COMPLETED
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import tkinter.font as font


class Scanner:
    def __init__(self) -> None:
        self.pdf_path = None
        self.writer = None
        self.highlights_by_color = {}
        self.pages_by_color = {}

    def rgb2color(self, rgb):
        """translate RGB to color name

        Args:
            rgb (str): rgb
        Returns:
            str: color name
        """
        color_names = {
            (0.9804, 0.9804, 0.0): "Yellow",
            (0.5647, 1.0, 0.5647): "Green",
            (0.0, 0.502, 1.0): "Blue",
            (1.0, 0.5647, 1.0): "Pink",
            (0.7294, 0.3333, 1.0): "Purple",
            (1.0, 0.5647, 0.5647): "Red",
            (1.0, 0.8471, 0.5647): "Orange",
            (0.8471, 0.8471, 0.8627): "Gray",
            (0.5647, 0.8471, 1.0): "Light Blue",
            (0.8471, 1.0, 0.8471): "Light Green",
            (0.8471, 0.5647, 0.0): "Brown"
        }
        closest_color = min(color_names.keys(), key=lambda c: sum((a-b)**2 for a, b in zip(c, rgb)))
        return color_names[closest_color]


    def scan_pdf(self, pdf_path, writer):
        """scan highlight in pdf file

        Args:
            pdf_path (_type_): _description_
            writer (_type_): _description_
        """
        self.pdf_path = pdf_path
        self.writer = writer
        self.highlights_by_color.clear()
        self.pages_by_color.clear()

        doc = fitz.open(self.pdf_path)

        try:
            for page_num, page in enumerate(doc):
                annotations = page.annots()
                if annotations:
                    for annot in annotations:
                        if annot.type[0] == 8:  
                            color = annot.colors['stroke']  
                            highlight = page.get_text("text", clip=annot.rect).strip()
                            if color not in self.highlights_by_color:
                                self.highlights_by_color[color] = []
                                self.pages_by_color[color] = []
                            if highlight in self.highlights_by_color[color]: # 去重
                                continue
                            self.highlights_by_color[color].append(highlight)
                            self.pages_by_color[color].append(page_num + 1)
        except Exception as e:
            print(e)
            print("some err occured.")
                

class ScannerGui(Scanner):
    """PDF Scanner Gui Class

    Args:
        Scanner (Scanner): Scanner class
    """
    def __init__(self, name, size, output_file, translator, excel_operator, youdao_worker) -> None:
        super().__init__()
        self.window = tk.Tk()
        self.name = name
        self.size = size
        self.output_file = output_file
        self.dir_path = None
        self.output_path = None
        self.api_key = ""
        self.api_secret = ""
        self.youdao_cookie_entry = ""
        self.selected_dictionary_entry = []
        self.translator = translator
        self.excel_operator = excel_operator
        self.youdao_worker = youdao_worker


    def run(self):
        self.window.title(self.name)
        self.window.geometry(self.size)

        global selected_directory,selected_dictionary, btn_start_scan, output_display, output_file_path_var, btn_open_file, trans_label, trans_label_var
        global translate_words, youdao_wordbook_check_var, output_words_excel, output_wordbook
        
        # 目录选择和显示
        selected_directory = tk.StringVar(self.window)
        tk.Entry(self.window, textvariable=selected_directory).grid(row=0, column=0, padx=(10,2),columnspan=3, pady=5, sticky='ew')
        tk.Button(self.window, text="选择文件", command=self.select_file).grid(row=0, column=4, padx=(0, 0), pady=5, sticky='ew')
        tk.Button(self.window, text="选择目录", command=self.select_directory).grid(row=0, column=5, padx=(0, 5), pady=5, sticky='ew')

        font_1 = font.Font(family='Arial', size=13, weight='bold')
        font_2 = font.Font(weight='bold')

        # 功能选项
        options_selected_Frame = tk.LabelFrame(self.window, text="功能选项", borderwidth=0, font=font_1)
        options_selected_Frame.grid(row=1, rowspan=1, column=0, padx=10, pady=5, sticky='ew')
        
        translate_words = tk.BooleanVar(value=True)
        translate_words_box = tk.Checkbutton(options_selected_Frame, text="翻译单词", variable=translate_words)
        translate_words_box.grid(row=0, column=0, sticky='w')
        
        youdao_wordbook_check_var = tk.BooleanVar(value=False)
        tk.Checkbutton(options_selected_Frame, text="添加单词本", variable=youdao_wordbook_check_var).grid(row=1, column=0, sticky='w')

        # 生成文件
        output_selected_Frame = tk.LabelFrame(self.window, text="生成文件", borderwidth=0, font=font_1)
        output_selected_Frame.grid(row=1,rowspan=1, column=1, padx=25, pady=5, sticky='ew')
        
        output_words_excel = tk.BooleanVar(value=True)
        output_words_box = tk.Checkbutton(output_selected_Frame, text="单词Excel", variable=output_words_excel)
        output_words_box.grid(row=0, column=1, sticky='w')

        output_wordbook = tk.BooleanVar(value=False)
        output_wordbook_box = tk.Checkbutton(output_selected_Frame, text="单词本xml", variable=output_wordbook)
        output_wordbook_box.grid(row=1, column=1, sticky='w')
        
        # 事件绑定
        def select_all_hook(source_var, *args):
            if source_var.get():
                for arg in args:
                    arg.set(source_var.get())  
                    
        def translate_wrods_hook(translate_words, output_words_excel, output_wordbook):
            if not translate_words.get():
                output_wordbook.set(translate_words.get())    
            output_words_excel.set(translate_words.get()) 
            
        def output_words_hook(output_words_excel, translate_words, output_wordbook):
            if not output_words_excel.get():
                output_wordbook.set(output_words_excel.get())
            translate_words.set(output_words_excel.get())            
            
        # 选项之间的关联性
        output_words_box.config(command=lambda: output_words_hook(output_words_excel, translate_words, output_wordbook))
        output_wordbook_box.config(command=lambda: select_all_hook(output_wordbook, translate_words, output_words_excel))
        translate_words_box.config(command=lambda: translate_wrods_hook(translate_words, output_words_excel, output_wordbook))
    
        # 导入配置
        btn_clear_output = tk.Button(self.window, text="添  加\n配  置", command=self.add_config)
        btn_clear_output.grid(row=1,rowspan=1, column=4, columnspan=1, padx=(0, 0), pady=(20,5), sticky='ew')

        # 开始扫描
        btn_start_scan = tk.Button(self.window, text="扫  描\nScan", command=self.start_scan, state=tk.DISABLED, font=font_2)
        btn_start_scan.grid(row=1, rowspan=1, column=5, columnspan=1, padx=(0, 5), pady=(20, 5), sticky='ew')

        # 带有垂直和水平滚动条的输出显示区域
        frame_output = tk.Frame(self.window)
        frame_output.grid(row=3, column=0, columnspan=6, padx=10, pady=5, sticky='nsew')

        v_scrollbar = tk.Scrollbar(frame_output, orient=tk.VERTICAL)
        h_scrollbar = tk.Scrollbar(frame_output, orient=tk.HORIZONTAL)
        
        output_display = tk.Text(frame_output, wrap=tk.NONE, yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set, height=18)
        v_scrollbar.config(command=output_display.yview)
        h_scrollbar.config(command=output_display.xview)
        
        output_display.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        frame_output.grid_columnconfigure(0, weight=1)
        frame_output.grid_rowconfigure(0, weight=1)

        # 输出文件路径和打开按钮
        output_file_path_var = tk.StringVar(self.window)
        tk.Entry(self.window, textvariable=output_file_path_var).grid(row=4, column=0, columnspan=5, padx=(10,2), pady=2, sticky='ew')
        btn_open_file = tk.Button(self.window, text="打开", command=self.open_output_file)
        btn_open_file.grid(row=4, column=5, padx=(1, 5), pady=5, sticky='ew')
        
        # 底部信息显示
        trans_label_var = tk.StringVar(self.window)
        trans_label = tk.Label(self.window, textvariable=trans_label_var, anchor=tk.W)
        trans_label.grid(row=5, column=0, columnspan=5, padx=(10,0), pady=5, sticky='ew')
 
        # 配置网格以适当地展开
        self.window.grid_columnconfigure(0, weight=1)
        self.window.grid_columnconfigure(1, weight=1)
        self.window.grid_columnconfigure(2, weight=1)
        self.window.mainloop()


    def check_path(self, path):
        """check selectd path

        Args:
            path (str): selected path

        Returns:
            bool: True or False
        """
        if not os.path.exists(path):
            messagebox.showwarning("警告", "路径不存在!")
            file_path = filedialog.askopenfile()
            if file_path:
                output_file_path_var.set(file_path.name)
                self.btn_change("normal")
                return True
            else:
                return False
        if not path.endswith(('.xls', '.xlsx')):
            messagebox.showwarning("警告", "请选择一个有效的Excel文件!")
            self.btn_change("normal")
            return False
        return True


    def select_file(self):
        """select file button
        """
        file_name = filedialog.askopenfile()
        if file_name:
            selected_directory.set(file_name.name)
            btn_start_scan.config(state=tk.NORMAL)


    def select_directory(self):
        """select directory button
        """
        dir_path = filedialog.askdirectory()
        if dir_path:
            selected_directory.set(dir_path)
            btn_start_scan.config(state=tk.NORMAL)


    def select_dictionary(self):
        """select dictionary button
        """
        dictionary_name = filedialog.askopenfile()
        if dictionary_name:
            selected_dictionary.delete(0, tk.END)
            selected_dictionary.insert(0, dictionary_name.name)


    def start_scan(self):
        """start scan button
        """
        if not selected_directory.get():
            messagebox.showwarning("Warning", "请先选择对应的文件或目录")
            return
        btn_start_scan.config(state=tk.DISABLED)
        output_display.delete(1.0, tk.END)  # clear output
        output_display.insert(tk.END, f"[Start] ***开始扫描*** \n")
        output_display.insert(tk.END, f"[Info] 请静侯,再静侯......\n\n")
        if os.path.isfile(selected_directory.get()):
            target = self.scan_file
        elif os.path.isdir(selected_directory.get()):
            target = self.scan_directory
        threading.Thread(target=target).start()


    def scan_file(self):
        """scan pdf file hightlight info and extract it into excel
        """
        file_path = selected_directory.get()
        self.dir_path = os.path.dirname(file_path)
        self.output_path= self.dir_path + "/" + self.output_file
        if not file_path.endswith('.pdf'):
            output_display.insert(tk.END, "[Error] 请选择PDF文件.\n")
            btn_start_scan.config(state=tk.NORMAL)
            return
        sheet_name = os.path.splitext(os.path.basename(file_path))[0][:31]
        mode='w'
        if os.path.exists(self.output_path):
            try:
                book = load_workbook(filename=self.output_path)
                if sheet_name in book.sheetnames:
                    if len(book.sheetnames)>1:
                        del book[sheet_name]
                        book.save(self.output_path)
                        mode='a'
                    else:
                        os.remove(self.output_path)
                        mode='w'
                else:
                    mode='a'
            except Exception as e:
                print("[ERROR] Error occurred: ", e)
        # 生成Excel
        self.btn_change("disable")
        if output_words_excel.get():
            if not self.excel_operator.generate(mytype="file", mode=mode, file_path=file_path, sheet_name=sheet_name):
                output_display.insert(tk.END, f"[Error] 处理Excel失败.\n")
            
        if not youdao_wordbook_check_var.get() and not translate_words.get() and not output_words_excel.get() and not output_wordbook.get():
            output_display.insert(tk.END, f"[Error] 至少要选一个功能, 不然啥也干不了呀宝贝:)\n\n")
            
        # 添加到单词本
        if youdao_wordbook_check_var.get(): 
            self.youdao_worker.add2youdao_wordbook(file_path=file_path)
            
        self.btn_change("normal")


    def scan_directory(self):
        """scan pdf files in directory and extract it into excel
        """
        dir_path = selected_directory.get()
        os.chdir(dir_path) 
        pdf_files = glob.glob('*.pdf')
        if not pdf_files:
            output_display.insert(tk.END, "[Error] 选定的目录中未发现PDF文件.\n")
            btn_start_scan.config(state=tk.NORMAL)
            return
        self.output_path = os.path.join(dir_path, self.output_file)
        
        # 生成Excel
        self.btn_change("disable")
        if output_words_excel.get():
            if not self.excel_operator.generate(mytype="directory", pdf_files=pdf_files):
                output_display.insert(tk.END, f"[Error] 处理Excel失败.\n")
                
        if not youdao_wordbook_check_var.get() and not translate_words.get() and not output_words_excel.get() and not output_wordbook.get():
            output_display.insert(tk.END, f"[Error] 至少要选一个功能, 不然啥也干不了呀宝贝:)\n\n")

        # 添加到单词本
        if youdao_wordbook_check_var.get():
            for pdf_file in pdf_files:
                self.youdao_worker.add2youdao_wordbook(file_path=pdf_file)
        self.btn_change("normal")

        
        
    def add_config(self):
        """add config to pdfScanner
        """
        self.new_window = tk.Toplevel(self.window)
        self.new_window.title("添加配置")
        
        title_font = font.Font(family='Arial', size=13, weight='bold')

        # 百度翻译api key-secret
        options_selected_Frame = tk.LabelFrame(self.new_window, text="百度翻译配置", borderwidth=0, font=title_font)
        options_selected_Frame.grid(row=1, rowspan=1, column=0, padx=10, pady=5, sticky='ew')
        
        tk.Label(options_selected_Frame, text="key:").grid(row=0, column=0, padx=10, pady=0, sticky=tk.W)
        baidu_api_key = tk.Entry(options_selected_Frame)
        baidu_api_key.grid(row=0, column=1, padx=10, pady=0)
        
        tk.Label(options_selected_Frame, text="secret:").grid(row=1, column=0, padx=10, pady=0, sticky=tk.W)
        baidu_api_secret = tk.Entry(options_selected_Frame)
        baidu_api_secret.grid(row=1, column=1, padx=10, pady=0)
        
        if self.translator.trans_id_pool:
            for i in self.translator.trans_id_pool.keys():
                self.api_key = i
                self.api_secret = self.translator.trans_id_pool[i]
        baidu_api_key.insert(0, self.api_key)
        baidu_api_secret.insert(0, self.api_secret)
        
        # 添加本地词典
        options_local_selected_Frame = tk.LabelFrame(self.new_window, text="本地词典", borderwidth=0, font=title_font)
        options_local_selected_Frame.grid(row=2, rowspan=1, column=0, padx=10, pady=5, sticky='ew')
        
        global selected_dictionary
        tk.Button(options_local_selected_Frame, text="选择", command=self.select_dictionary).grid(row=0, column=0, padx=(10,5), pady=0, sticky=tk.W)
        selected_dictionary = tk.Entry(options_local_selected_Frame)
        selected_dictionary.grid(row=0, column=1, padx=0, pady=0, sticky=tk.W)
        
        if self.translator.book and not self.selected_dictionary_entry:
            self.selected_dictionary_entry = self.translator.book[0]
        selected_dictionary.insert(0, self.selected_dictionary_entry)
        
        # 网易有道词典cookie
        options_cookie_selected_Frame = tk.LabelFrame(self.new_window, text="有道词典配置", borderwidth=0, font=title_font)
        options_cookie_selected_Frame.grid(row=3, rowspan=1, column=0, padx=10, pady=10, sticky='ew')
        
        tk.Label(options_cookie_selected_Frame, text="cookie:").grid(row=0, column=0, padx=10, pady=0, sticky=tk.W)
        youdao_cookie = tk.Entry(options_cookie_selected_Frame)
        youdao_cookie.grid(row=0, column=1, padx=10, pady=0)
        
        if self.youdao_worker.cookie and not self.youdao_cookie_entry:
            self.youdao_cookie_entry = self.youdao_worker.cookie

        youdao_cookie.insert(0, self.youdao_cookie_entry)
        
        confirm_button = tk.Button(self.new_window, text="确认", command=lambda: self.sub_window_sommit(str(baidu_api_key.get()),
                                                                                                      str(baidu_api_secret.get()),
                                                                                                      str(youdao_cookie.get()),
                                                                                                      str(selected_dictionary.get())
                                                                                                      ),)
        confirm_button.grid(row=4, column=0, columnspan=2, pady=20)
        
        # 设置新窗口位置
        self.calculate_sub_window_pos()


    def sub_window_sommit(self, baidu_api_key, baidu_api_secret, cookie_entry, dictionary):
        """add config window confirm button

        Args:
            baidu_api_key (str): baidu_api_key
            baidu_api_secret (str): baidu_api_secret
            cookie_entry (str): cookie_entry
            dictionary (str): selected dictionary
        """
        message = ""
        if baidu_api_key and baidu_api_secret:
            self.api_key = baidu_api_key
            self.api_secret = baidu_api_secret
            self.translator.trans_id_pool[baidu_api_key] = baidu_api_secret
        else:
            message = "百度api或secret若为空, 则无法进行翻译!\n"
        if dictionary:
            self.selected_dictionary_entry = dictionary
            self.translator.book.append(dictionary)
        if cookie_entry:
            self.youdao_cookie_entry = cookie_entry
            self.youdao_worker.cookie = cookie_entry
        else:
            message += "有道翻译cookie若为空, 则无法添加单词本!\n"
        if message:
            messagebox.showwarning("Warning", message)
        self.new_window.destroy()
        
        
    def btn_change(self, state):
        """start scan button click change

        Args:
            state (str): disable or norma
        """
        if state == "disable":
            btn_start_scan.config(state=tk.DISABLED)
            btn_open_file.config(state=tk.DISABLED) 
        elif state == "normal":
            btn_start_scan.config(state=tk.NORMAL)
            btn_open_file.config(state=tk.NORMAL) 


    def calculate_sub_window_pos(self):
        """calculate sub window position
        """
        self.new_window.update_idletasks() 
        width = self.new_window.winfo_width()
        x = self.window.winfo_x() + (self.window.winfo_width() - width) // 2  
        y = self.window.winfo_y() + self.window.winfo_height() // 5  
        self.new_window.geometry(f"+{x}+{y}")


    def open_output_file(self):
        """open output excel file 
        """
        path = output_file_path_var.get().strip()
        if not os.path.exists(path):
            messagebox.showwarning("警告", "路径不存在!")
            file_path = filedialog.askopenfile()
            if file_path:
                output_file_path_var.set(file_path.name)
            return
        if not path.endswith(('.xls', '.xlsx')):
            messagebox.showwarning("警告", "请选择一个有效的Excel文件!")
            return
        file_path = output_file_path_var.get().strip()
        if sys.platform.startswith("darwin"):
            os.system(f'open "{file_path}"')
        elif sys.platform.startswith("win32"):
            os.system(f'start "" "{file_path}"')


class Translator:
    """Translator Class
    """
    def __init__(self, gui_obj, trans_id_pool, base_url,words_book_url, sleep_time, book=[]) -> None:
        self.gui_obj = gui_obj
        self.trans_id_pool = trans_id_pool
        self.BASE_URL = base_url
        self.words_book_url = words_book_url
        self.sleep_time = sleep_time
        self.book = book
        self.dictionary = {}    # 加载的词典
        
      
    def load_translate_books(self):
        """load local dictionary

        Returns:
            None
        """
        n = 0
        for book_item in self.book: 
            my_dictionary = pd.read_csv(book_item, sep='⬄', header=0, names=['word', 'interpretation'])
            def check(series):
                return series['word'].strip()
            my_dictionary['word'] = my_dictionary.apply(check, axis=1)
            my_dictionary.set_index('word', inplace=True)
            self.dictionary[n] = my_dictionary
            n += 1


    def start(self):
        """start translater
        """
        # output_display.delete(1.0, tk.END)  # clear output
        trans_label_var.set("加载词典中......")
        self.load_translate_books()
        threading.Thread(target=self.start_translate).start()
        
        
    def start_translate(self):
        """translate words in excel
        """
        path = output_file_path_var.get().strip()
        if not self.gui_obj.check_path(path): 
            return

        self.gui_obj.btn_change("disable")
        trans_label_var.set("正在拼命翻译中, 请耐心等待......")
        output_display.insert(tk.END, f"[Info] 翻译开始, 请等待一段时间...\n")

        # 对Excel中的单词翻译
        self.gui_obj.excel_operator.execl_deal2translat(path=path)

        output_display.insert(tk.END, f"[Success] 翻译完成!\n")
        current_xview = output_display.xview()
        output_display.see(tk.END)
        output_display.xview_moveto(current_xview[0])
        
        # 生成有道单词本
        self.gui_obj.youdao_worker.generate_youdao_workbook(path=path)
        
        trans_label_var.set("翻译完成!")
        self.gui_obj.btn_change("normal")


    def get_trans_id(self):
        key = random.choice(list(self.trans_id_pool))
        value = self.trans_id_pool[key]
        return (str(key), value)


    def trans_row(self, text, ws, row, col):
        """translate one cell in Excel

        Args:
            text (str): content in excel cell
            ws (w): worksheet
            row (int): row
            col (int): column
        """
        # 进行翻译
        ret = self.translate_local(str(text))
        if ret:
            translation = str(ret)
        else:
            translation = self.translate_baidu_api(str(text))
            
        # 处理翻译内容
        if len(translation) > 150:
            translation = translation[:150]
            
        # if "sup>" in translation:
        #     translation = translation.split("sup>")[2]       
        # if " ➞ " in translation:
        #     translation = translation.split("➞")[0]
        # # if "：" in translation:
        # #     translation = translation.split("：")[0]
        # # if ":" in translation:
        # #     translation = translation.split(":")[0]
        # if "/" in translation:
        #     translation = translation.split("/")[0]
            
        # 创建有道词典单词本xml对象
        if output_wordbook.get():
            self.gui_obj.youdao_worker.create_wb_xml(str(text), translation)
            
        current_xview = output_display.xview()
        output_display.see(tk.END)
        output_display.xview_moveto(current_xview[0])
        if translation:
            ws.cell(row=row, column=col+1).value = translation
        time.sleep(float(self.sleep_time))


    def translate_baidu_api(self, text):
        """use baidu translate api 

        Args:
            text (str): word
        Returns:
            str: translate result
        """
        APP_ID, SECRET_KEY = self.get_trans_id()
        salt = 'random_salt'
        sign = APP_ID + text + salt + SECRET_KEY
        m = hashlib.md5()
        m.update(sign.encode('utf-8'))
        sign = m.hexdigest()
        params = {
            'q': text,
            'from': 'en',
            'to': 'zh',
            'appid': APP_ID,
            'salt': salt,
            'sign': sign
        }
        response = requests.get(self.BASE_URL, params=params)
        result = response.json()
        for i in range(3):
            try:
                dst = str(result['trans_result'][0]['dst'])
                output_display.insert(tk.END, f"[Info] {text}: {dst}\n")
                return dst
            except KeyError:
                if i < 2:
                    time.sleep(0.2)
        print('翻译失败:', result)
        return None


    def translate_local(self, word):
        """use local dictionary to translate

        Args:
            word (str): word
        Returns:
            str: translated word 
        """
        for i in self.dictionary.keys():
            if word in self.dictionary[i].index.values:
                result = self.dictionary[i].loc[word].values.tolist()[0]
                output_display.insert(tk.END, f"[Info] {word}: {result}\n")
                return result
            else:
                continue
        return None


class Youdao_worker:
    """youdao wordbook class
    """
    def __init__(self, gui_obj, words_book_url, tag_name, cookie="") -> None:
        self.gui_obj = gui_obj
        self.words_book_url = words_book_url
        self.tag_name = tag_name
        self.youdao_book = ET.Element('youdao_wordbook.xml')
        self.cookie = cookie

        
    def add2youdao_wordbook(self, file_path):
        """add words to youdao wordbook

        Args:
            file_path (str): pdf file's absolutely path
        """
        file_name = os.path.basename(file_path)
        if not self.cookie:
            output_display.insert(tk.END, f"[Error] 添加到单词本失败!\n")
            output_display.insert(tk.END, f"[Error] 添加到单词本需要先添加cookie配置!\n\n")
            messagebox.showwarning("警告", "请先添加有道cookie的配置!")
            self.gui_obj.btn_change("normal")
            return
        
        self.gui_obj.scan_pdf(file_path, None)
        if self.gui_obj.highlights_by_color:
            t_poll = ThreadPoolExecutor(max_workers=16)
            thread_list = []
            output_display.insert(tk.END, f"[Info] 开始添加PDF文件中的单词到有道单词本...\n")
            output_display.insert(tk.END, f"[Info] PDF文件名: {file_name}\n")
            for color in self.gui_obj.highlights_by_color.keys():
                for word_hightlight in self.gui_obj.highlights_by_color[color]:
                    f = t_poll.submit(self.youdao_wordbook_request, str(word_hightlight).strip())
                    thread_list.append(f)
            wait(thread_list, return_when=ALL_COMPLETED)
            output_display.insert(tk.END, f"[Info] 完成! 当前PDF中的单词添加到单词本完成!\n\n")
            current_xview = output_display.xview()
            output_display.see(tk.END)
            output_display.xview_moveto(current_xview[0])
            
            
    def youdao_wordbook_request(self, word):
        """add word to youdao wordbook

        Args:
            word (str): wrod
        Returns:
            None
        """
        url = f"{self.words_book_url}{word}"
        headers = {
            'Cookie': self.cookie,
            'Host': 'dict.youdao.com',
            'Upgrade-Insecure-Requests': '1',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Accept': 'application/json, text/plain, */*',
            'Referer': 'https://dict.youdao.com',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_1_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'
        }
        for i in range(3):
            try:
                response = requests.get(url, headers=headers)
                data = response.json()
                if data['code'] == 0:
                    output_display.insert(tk.END, f"[Info] Success: 添加 {str(word)} 到单词本!\n")
                    current_xview = output_display.xview()
                    output_display.see(tk.END)
                    output_display.xview_moveto(current_xview[0])
                    return
                else:
                    continue
            except Exception as e:
                if i < 2:
                    time.sleep(0.2)
        output_display.insert(tk.END, f"[Error] Failed: 添加 {str(word)} 失败!\n")
        current_xview = output_display.xview()
        output_display.see(tk.END)
        output_display.xview_moveto(current_xview[0])
        return "Failed"
    

    def generate_youdao_workbook(self, path):
        # 生成有道词典单词本xml文件
        if output_wordbook.get():
            xml_str = ET.tostring(self.youdao_book, encoding='unicode', method='xml')
            xml_str = xml_str.replace('&lt;', '<').replace('&gt;', '>').replace("youdao_wordbook.xml", "wordbook")
            bookfile = "%s/youdao_wordbook.xml" % os.path.dirname(path)
            with open(bookfile, 'wb') as f:
                f.write(xml_str.encode('utf-8'))
            output_display.insert(tk.END, f"[Success] 单词本: {bookfile}\n")
            

    def create_wb_xml(self, word, trans):
        """generate youdao wordbook xml object

        Args:
            word (str): word
            trans (str): translate result
        """
        item = ET.SubElement(self.youdao_book, 'item')
        word_elem = ET.SubElement(item, 'word')
        word_elem.text = f'<![CDATA[{word}]]>'
        trans_elem = ET.SubElement(item, 'trans')
        trans_elem.text = f'<![CDATA[{trans}]]>'
        phonetic_elem = ET.SubElement(item, 'phonetic')
        phonetic_elem.text = '<![CDATA[]]>'
        tags_elem = ET.SubElement(item, 'tags')
        tags_elem.text = '<![CDATA[%s]]>' % self.tag_name
        progress_elem = ET.SubElement(item, 'progress')
        progress_elem.text = '0'


class Excel_operator:
    """Excel operator class
    """
    def __init__(self,gui_obj, view_workers, row_workers) -> None:
        self.gui_obj = gui_obj
        self.view_workers = view_workers
        self.row_workers = row_workers
        
        
    def generate(self, mytype, mode=None, file_path=None, pdf_files=None, sheet_name=None) -> bool:
        """generate Excel file to operate

        Args:
            mytype (str): file or directory
            mode (str, optional): a or w. Defaults to None.
            file_path (str, optional): file path. Defaults to None.
            pdf_files (list, optional): pdf files list. Defaults to None.
            sheet_name (str, optional): sheet name. Defaults to None.
        Returns:
            bool: success or failed
        """
        if file_path and mode and mytype == "file":
            with pd.ExcelWriter(self.gui_obj.output_path, engine='openpyxl', mode=mode) as writer:
                self.gui_obj.scan_pdf(file_path, writer)
                self.excel_scan2deal(sheet_name)
            self.excel_translate()
            return True
        elif pdf_files and mytype == "directory":
            with pd.ExcelWriter(self.gui_obj.output_path, engine='openpyxl') as writer:
                for pdf_file in pdf_files:
                    self.gui_obj.scan_pdf(pdf_file, writer)     
                    sheet_name = os.path.splitext(os.path.basename(self.gui_obj.pdf_path))[0][:31]
                    self.excel_scan2deal(sheet_name)
            self.excel_translate()
            return True
        else:
            return False
        
        
    def excel_scan2deal(self, sheet_name):
        """Convert data to DataFrame for Excel

        Args:
            sheet_name (str): Excel sheet name
        """      
        if self.gui_obj.highlights_by_color: 
            output_display.insert(tk.END, f"[Info] PDF文件扫描完成!\n")
            output_display.insert(tk.END, f"[Info] 开始处理: {self.gui_obj.pdf_path}\n")
            max_rows = max(len(texts) for texts in self.gui_obj.highlights_by_color.values())
            df = pd.DataFrame()
            for color in self.gui_obj.highlights_by_color.keys():
                col_name = self.gui_obj.rgb2color(color)
                df[col_name] = self.gui_obj.highlights_by_color[color] + [''] * (max_rows - len(self.gui_obj.highlights_by_color[color]))
                df[col_name + " Page"] = self.gui_obj.pages_by_color[color] + [''] * (max_rows - len(self.gui_obj.pages_by_color[color]))
            df.to_excel(self.gui_obj.writer, index=False, sheet_name=sheet_name)
        else:
            print(f"No highlighted text found in {self.gui_obj.pdf_path}.")
            output_display.insert(tk.END, f"[Warn] 无高亮: {self.gui_obj.pdf_path}\n")
            

    def excel_translate(self):
        """deal excel sheet and translate
        """
        self.excel_sort_sheets()
        output_display.insert(tk.END, f"[Info] 处理Excel归类完成!\n")
        output_display.insert(tk.END, f"[Info] 生成文件: {self.gui_obj.output_path}.\n\n")
        output_file_path_var.set(self.gui_obj.output_path)
        
        # 是否翻译
        if translate_words.get():
            self.gui_obj.translator.start()
            
        btn_start_scan.config(state=tk.NORMAL) 
        current_xview = output_display.xview()
        output_display.see(tk.END)
        output_display.xview_moveto(current_xview[0])
        

    def excel_sort_sheets(self):
        """sort sheet by sheetname
        """
        sheet = load_workbook(self.gui_obj.output_path)
        sorted_sheet_names = sorted(sheet.sheetnames)
        print(sorted_sheet_names)
        for idx, sheet_name in enumerate(sorted_sheet_names):
            sheet[sheet_name].index = idx
        sheet.save(self.gui_obj.output_path)
    
    
    def execl_deal2translat(self, path):
        wb = openpyxl.load_workbook(path)
        t_poll = ThreadPoolExecutor(max_workers=self.view_workers)
        thread_list = []
        for worksheet_name in wb.sheetnames:
            output_display.insert(tk.END, f"[Info] 开始翻译: {worksheet_name}\n")
            ws = wb[worksheet_name]
            if ws.max_column % 3 != 0:
                f = t_poll.submit(self.insert_translation_columns, ws)
                thread_list.append(f)   
        wait(thread_list, return_when=ALL_COMPLETED)
        # 保存为Excel文件
        wb.save(path)
        
        
    def insert_translation_columns(self, ws):
        """insert translated info to specific cell in column

        Args:
            ws (str): worksheet name
        """
        max_col = ws.max_column
        for col in range(1, 2*max_col+1, 3):
            ws.insert_cols(col+1)
            executor = ThreadPoolExecutor(max_workers=self.row_workers)
            t_list = []
            # 从第二行开始，翻译并填充翻译列的内容
            for row in range(2, ws.max_row + 1):
                text = ws.cell(row=row, column=col).value
                if text:
                    f = executor.submit(self.gui_obj.translator.trans_row, text, ws, row, col)
                    t_list.append(f)
            wait(t_list, return_when=ALL_COMPLETED)
        
        
def main():
    # 网页有道翻译的cookie
    youdao_cookie = ''

    # 百度翻译的api, 现只可以填入一个key-secret
    id_pool = {
        # '填入你的key': '填入你的secret',
    }
    
    # 本地字典的路径
    translate_book = [
        "./Dictionary/英汉大词典_del_ipa_edited.txt",
    ]
    
    youdao_worker = Youdao_worker(
        gui_obj = None,
        tag_name="mytest",            # 生成的有道单词本xml中的tag
        words_book_url='https://dict.youdao.com/wordbook/webapi/v2/ajax/add?lan=en&word=',  # 有道单词本api
        cookie=youdao_cookie,
    )
    
    excel_operator = Excel_operator(
        gui_obj = None,
        view_workers=1,
        row_workers=6,
    )
    
    translator = Translator(
        gui_obj = None,
        trans_id_pool=id_pool,
        base_url='https://fanyi-api.baidu.com/api/trans/vip/translate',                     # 百度翻译api
        words_book_url='https://dict.youdao.com/wordbook/webapi/v2/ajax/add?lan=en&word=',  # 有道单词本api
        sleep_time='0.1',
        book=translate_book,
    )
    
    scanner_gui = ScannerGui(
        name="pdfScanner v1.5.2",
        size="450x450",
        output_file="output.xlsx",    # 生成excel文件的名字
        translator=translator,
        excel_operator=excel_operator,
        youdao_worker=youdao_worker,
    )
    
    youdao_worker.gui_obj = scanner_gui
    excel_operator.gui_obj = scanner_gui
    translator.gui_obj = scanner_gui
    
    scanner_gui.run()

if __name__ == "__main__":
    main()
